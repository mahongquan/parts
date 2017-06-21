# -*- coding: utf-8 -*-
from lxml import etree as ET
from io import BytesIO,StringIO
import datetime
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from genDoc.genData import genTest,genjmd,genTestR
import os
from mysite.settings import MEDIA_ROOT
import logging
def getElement(chanels,first):
    print(first)
    eles=first.split("(")
    ele=eles[0]#2C
    if ele[0]=="2":
        chanels.append("L"+ele[1])
        chanels.append("H"+ele[1])
    else:
        ele_set=eles[1][:-1]#移去括号
        print(ele_set)
        if ele_set[0]=="高":
            chanels.append("H"+ele[1])
        else:
            chanels.append("L"+ele[1])
    pass    
def getchannels(peizhi):
    print(peizhi)
    elements=peizhi.split("+")
    chanels=[]
    first=elements[0]
    getElement(chanels,first)
    if len(elements)==1:#单元素
        pass
    else:#two elements
        second=elements[1]
        getElement(chanels,second)
    return chanels
def setCell(table,row,col,value):
    print(row,col,value)
    table.cell(row = row, column = col).value=value
def getCell(table,row,col):
    return table.cell(row = row, column = col).value
def genJiaozhunO(c,fn):
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0] 
    setCell(table,16,8,c.yiqibh)#setCell(table,10,8).Value =c.yiqixinghao    setCell(table,11,8).Value =c.yiqixinghao
    setCell(table,18,8,c.yonghu)#setCell(table,18,8).Value =c.yonghu
    setCell(table,20,8,c.addr)# setCell(table,18,8).Value =c.yonghu
    d=c.yujifahuo_date
    d1=d+datetime.timedelta(-1)
    setCell(table,32,8,str(d1.year))#setCell(table,32,8).FormulaR1C1 =d1.year#年
    setCell(table,32,12,str(d1.month))#setCell(table,32,12).Value =d1.month#月
    setCell(table,32,16,str(d1.day))#setCell(table,32,16).Value =d1.day#日
    d2=d+datetime.timedelta(364)
    setCell(table,35,8,str(d2.year))#setCell(table,35,8).Value =d2.year#年
    setCell(table,35,12,str(d2.month))#setCell(table,35,12).Value =d2.month#月
    setCell(table,35,16,str(d2.day))#setCell(table,35,16).Value =d2.day#日
    # #page 2
    table=xlBook.worksheets[1] 
    #dd="  地点（LOCATION）： "+c.yonghu
    setCell(table,22,3,"地点"+"("+"LOCATION"+")"+":"+"    "+c.yonghu)# setCell(table,22,3).Value=dd
    # #page 3
    table=xlBook.worksheets[2] 
    eles=[]
    for i in [8,15]:
        print("getCell",13,i)
        eles.append(getCell(table,13,i))#.Value)
    stds=[]
    for i in [8,15]:
        stds.append(getCell(table,14,i))#setCell(table,14,i).Value)
    (tests,errs)=genTestR(eles,stds)
    setCell(table,15,8,tests[0]) # setCell(table,15,8).Value=tests[0]
    setCell(table,15,15,tests[1]) # setCell(table,15,15).Value=tests[1]
    setCell(table,16,8,errs[0]) # setCell(table,16,8).Value=errs[0]
    setCell(table,16,15,errs[1]) # setCell(table,16,15).Value=errs[1]
    # #jmd
    cave=0.0134
    crsd=0.74/100
    (rs,ave_str,rsd_str)=genjmd(cave,crsd)
    cjmd_str=",".join(rs)
    setCell(table,20,3,"   测量值(O/%):" +cjmd_str) # setCell(table,20,3).Value="   测量值(O/%):" +cjmd_str
    setCell(table,21,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%")) # setCell(table,21,3).Value="   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%")
    return save_virtual_workbook(xlBook)
def genJiaozhunOH(c,fn):
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0] 
    setCell(table,10,8,c.yiqixinghao)
    setCell(table,16,8,c.yiqibh)
    setCell(table,18,8,c.yonghu)
    setCell(table,20,8,c.addr)# setCell(table,18,8).Value =c.yonghu
    d=c.yujifahuo_date
    d1=d+datetime.timedelta(-1)
    setCell(table,32,8,str(d1.year))#年
    setCell(table,32,12,str(d1.month))#月
    setCell(table,32,16,str(d1.day))#日
    d2=d+datetime.timedelta(364)
    setCell(table,35,8,str(d2.year))#年
    setCell(table,35,12,str(d2.month))#月
    setCell(table,35,16,str(d2.day))#日
    #page 2
    table=xlBook.worksheets[1] 
    #dd="  地点（LOCATION）： "+c.yonghu
    setCell(table,23,3,"  "+"地点"+"("+"LOCATION"+")"+":"+"    "+c.yonghu)# setCell(table,22,3).Value=dd#setCell(table,22,3,dd)setCell(table,23,3,dd)
    #page 3
    table=xlBook.worksheets[2] 
    eles=[]
    for i in [6,9,12,14,17,20]:
        eles.append(getCell(table,13,i))
    stds=[]
    for i in [6,9,12,14,17,20]:
        stds.append(getCell(table,14,i))
    print(eles,stds)
    (tests,errs)=genTestR(eles,stds)
    tmp=0
    for i in [6,9,12,14,17,20]:
        setCell(table,15,i,tests[tmp])
        tmp +=1
    tmp=0
    for i in [6,9,12,14,17,20]:
        setCell(table,16,i,errs[tmp])
        tmp +=1
    #jmd
    cave=0.1886
    crsd=1.27/100
    (rs,ave_str,rsd_str)=genjmd(cave,crsd)
    cjmd_str=",".join(rs)
    setCell(table,20,3,"   测量值(O/%):" +cjmd_str)
    setCell(table,21,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    sjmd_str=" 0.0721,0.0725,0.0723,0.0727,0.0730,0.0725,0.0725"
    save=0.00181
    srsd=1.1/100
    (rs,ave_str,rsd_str)=genjmd(save,srsd)
    sjmd_str=",".join(rs)
    setCell(table,22,3,"   测量值(H/%):" + sjmd_str)
    setCell(table,23,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    return save_virtual_workbook(xlBook)
def genJiaozhunN(c,fn):
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0] 
    setCell(table,10,8,c.yiqixinghao)
    setCell(table,16,8,c.yiqibh)
    setCell(table,18,8,c.yonghu)
    setCell(table,20,8,c.addr)# setCell(table,18,8).Value =c.yonghu
    d=c.yujifahuo_date
    d1=d+datetime.timedelta(-1)
    setCell(table,32,8,str(d1.year))#年
    setCell(table,32,12,str(d1.month))#月
    setCell(table,32,16,str(d1.day))#日
    d2=d+datetime.timedelta(364)
    setCell(table,35,8,str(d2.year))#年
    setCell(table,35,12,str(d2.month))#月
    setCell(table,35,16,str(d2.day))#日
    #page 2
    table=xlBook.worksheets[1] 
    dd="  地点（LOCATION）： "+c.yonghu
    setCell(table,22,3,"地点"+"("+"LOCATION"+")"+":"+"    "+c.yonghu)# setCell(table,22,3).Value=dd#setCell(table,22,3,dd)
    #page 3
    table=xlBook.worksheets[2] 
    eles=[]
    for i in [11,18]:
        eles.append(getCell(table,13,i))
    stds=[]
    for i in [11,18]:
        stds.append(getCell(table,14,i))
    print(eles,stds)
    (tests,errs)=genTestR(eles,stds)
    tmp=0
    for i in [11,18]:
        setCell(table,15,i,tests[tmp])
        tmp +=1
    tmp=0
    for i in [11,18]:
        setCell(table,16,i,errs[tmp])
        tmp +=1
    #jmd
    save=0.0084
    srsd=1.5/100
    (rs,ave_str,rsd_str)=genjmd(save,srsd)
    sjmd_str=",".join(rs)
    setCell(table,22,3,"   测量值(N/%):" + sjmd_str)
    setCell(table,23,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    return save_virtual_workbook(xlBook)
def genJiaozhunON(c,fn):
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0] 
    setCell(table,10,8,c.yiqixinghao)# setCell(table,11,8).Value =c.yiqixinghao
    setCell(table,16,8,c.yiqibh)# setCell(table,16,8).Value =c.yiqibh
    setCell(table,18,8,c.yonghu)# setCell(table,18,8).Value =c.yonghu
    setCell(table,20,8,c.addr)# setCell(table,18,8).Value =c.yonghu
    d=c.yujifahuo_date
    d1=d+datetime.timedelta(-1)
    setCell(table,32,8,str(d1.year))# setCell(table,32,8).FormulaR1C1 =d1.year#年
    setCell(table,32,12,str(d1.month))# setCell(table,32,12).Value =d1.month#月
    setCell(table,32,16,str(d1.day))# setCell(table,32,16).Value =d1.day#日
    d2=d+datetime.timedelta(364)
    setCell(table,35,8,str(d2.year))# setCell(table,35,8).Value =d2.year#年
    setCell(table,35,12,str(d2.month))# setCell(table,35,12).Value =d2.month#月
    setCell(table,35,16,str(d2.day))# setCell(table,35,16).Value =d2.day#日
    # #page 2
    table=xlBook.worksheets[1] 
    dd="  地点（LOCATION）： "+c.yonghu
    setCell(table,22,3,"地点"+"("+"LOCATION"+")"+":"+"    "+c.yonghu)# setCell(table,22,3).Value=ddsetCell(table,22,3,dd)# setCell(table,26,3,dd
    # #page 3
    table=xlBook.worksheets[2] 
    eles=[]
    for i in [8,11,15,18]:
        eles.append(getCell(table,13,i))
    stds=[]
    for i in [8,11,15,18]:
        stds.append(getCell(table,14,i))
    # print(eles,stds)
    (tests,errs)=genTestR(eles,stds)
    tmp=0
    for i in [8,11,15,18]:
        setCell(table,15,i,tests[tmp])
        tmp +=1
    tmp=0
    for i in [8,11,15,18]:
        setCell(table,16,i,errs[tmp])
        tmp +=1
    #jmd
    cave=0.0112
    crsd=0.8/100
    (rs,ave_str,rsd_str)=genjmd(cave,crsd)
    cjmd_str=",".join(rs)
    setCell(table,20,3,"   测量值(O/%):" +cjmd_str)
    setCell(table,21,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    sjmd_str=" 0.0721,0.0725,0.0723,0.0727,0.0730,0.0725,0.0725"
    save=0.0084
    srsd=1.3/100
    (rs,ave_str,rsd_str)=genjmd(save,srsd)
    sjmd_str=",".join(rs)
    setCell(table,22,3,"   测量值(N/%):" + sjmd_str)
    setCell(table,23,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    return save_virtual_workbook(xlBook)
def genJiaozhunCS(c,fn):  
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0]  
    #Worksheet = Worksheets[0]
    #table=Worksheet.find("{urn:schemas-microsoft-com:office:spreadsheet}Table")
    setCell(table,10,8,c.yiqixinghao)#setCell(table,10,8).Value =c.yiqixinghao
    setCell(table,16,8,c.yiqibh)# setCell(table,16,8).Value =c.yiqibh
    setCell(table,18,8,c.yonghu)# setCell(table,18,8).Value =c.yonghu
    setCell(table,20,8,c.addr)# setCell(table,18,8).Value =c.yonghu
    d=c.yujifahuo_date
    d1=d+datetime.timedelta(-1)
    setCell(table,32,8,str(d1.year))# setCell(table,32,8).FormulaR1C1 =d1.year#年
    setCell(table,32,12,str(d1.month))# setCell(table,32,12).Value =d1.month#月
    setCell(table,32,16,str(d1.day))# setCell(table,32,16).Value =d1.day#日
    d2=d+datetime.timedelta(364)
    setCell(table,35,8,str(d2.year))# setCell(table,35,8).Value =d2.year#年
    setCell(table,35,12,str(d2.month))# setCell(table,35,12).Value =d2.month#月
    setCell(table,35,16,str(d2.day))# setCell(table,35,16).Value =d2.day#日
    # #page 2
    #Worksheet = Worksheets[1]
    #table=Worksheet.find("{urn:schemas-microsoft-com:office:spreadsheet}Table")
    table=xlBook.worksheets[1]  
    #dd="  地点（LOCATION）： "+c.yonghu
    setCell(table,26,3,"地点"+"("+"LOCATION"+")"+":"+"    "+c.yonghu)# setCell(table,22,3).Value=ddsetCell(table,26,3,dd)# setCell(table,26,3,dd
    # #page 3
    #Worksheet = Worksheets[2]
    #table=Worksheet.find("{urn:schemas-microsoft-com:office:spreadsheet}Table")
    table=xlBook.worksheets[2]  
    eles=[]
    for i in range(6):
        eles.append(getCell(table,13,8+i*2))#setCell(table,13,8+i*2).Value)
    stds=[]
    for i in range(6):
        stds.append(getCell(table,14,8+i*2))#setCell(table,14,8+i*2).Value)
    (tests,errs)=genTest(eles,stds)
    for i in range(6):
        setCell(table,15,8+i*2,tests[i])#setCell(table,15,8+i*2).Value=tests[i]
    for i in range(6):
        setCell(table,16,8+i*2,errs[i])#setCell(table,16,8+i*2).Value=errs[i]
    # #jmd
    cave=0.0725
    crsd=0.3/100
    (rs,ave_str,rsd_str)=genjmd(cave,crsd)
    cjmd_str=",".join(rs)
    setCell(table,20,3,"   测量值(C/%):" +cjmd_str)# setCell(table,20,3).Value="   测量值(C/%):" +cjmd_str
    setCell(table,21,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))# setCell(table,21,3).Value="   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%")
    sjmd_str=" 0.0721,0.0725,0.0723,0.0727,0.0730,0.0725,0.0725"
    save=0.0723
    srsd=1.04/100
    (rs,ave_str,rsd_str)=genjmd(save,srsd)
    sjmd_str=",".join(rs)
    setCell(table,22,3,"   测量值(S/%):" + sjmd_str)# setCell(table,22,3).Value="   测量值(S/%):" + sjmd_str
    setCell(table,23,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))# setCell(table,23,3).Value="   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%")
    return save_virtual_workbook(xlBook)
def genShujubiao(c,fn):
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0] 
    channels=getchannels(c.channels)
    if c.yiqixinghao=="CS-2800":
        setCell(table,4,3,getCell(table,4,3)+"√") #setCell(table,4,3).Value =setCell(table,4,3).Value     +"√"
        if "LC" in channels:
            setCell(table,8,5,"√")
        if "LS" in channels:
            setCell(table,8,6,"√")
        if "HC" in channels:
            setCell(table,8,3,"√")
        if "HS" in channels:
            setCell(table,8,4,"√")
    elif c.yiqixinghao=="CS-3000":
        setCell(table,4,4,getCell(table,4,4)+"√") #setCell(table,4,3).Value =setCell(table,4,3).Value     +"√"
        setCell(table,8,3,"√")
        setCell(table,8,5,"√")
        setCell(table,8,6,"√")
    elif c.yiqixinghao=="O-3000":
        setCell(table,4,5,getCell(table,4,5)+"√")#setCell(table,4,5).Value =setCell(table,4,5).Value     +"√"
        setCell(table,11,3,"√")#setCell(table,11,3).Value ="√"
    elif c.yiqixinghao=="N-3000":
        setCell(table,4,6,getCell(table,4,6)+"√")#setCell(table,4,6).Value =setCell(table,4,6).Value     +"√"
    elif c.yiqixinghao=="ON-3000":
        setCell(table,5,3,getCell(table,5,3)+"√")#setCell(table,5,3).Value =setCell(table,5,3).Value     +"√"
        setCell(table,11,3,"√")#setCell(table,11,3).Value ="√"
        setCell(table,11,4,"√")#setCell(table,11,4).Value ="√"
    elif c.yiqixinghao=="ONH-3000":
        setCell(table,5,4,getCell(table,5,4)+"√")#setCell(table,5,4).Value =setCell(table,5,4).Value     +"√"
        setCell(table,11,3,"√")#setCell(table,11,3).Value ="√"
        setCell(table,11,4,"√")#setCell(table,11,4).Value ="√"
        setCell(table,11,5,"√")#setCell(table,11,3).Value ="√"
    elif c.yiqixinghao=="OH-3000":
        setCell(table,5,6,getCell(table,5,6)+"√")#setCell(table,5,6).Value =setCell(table,5,6).Value     +"√"
        setCell(table,11,3,"√")#setCell(table,11,3).Value ="√"
        setCell(table,11,5,"√")#setCell(table,11,5).Value ="√"
    setCell(table,2,6,"合同号："+c.hetongbh)#setCell(table,2,6).Value ="合同号："+c.hetongbh    
    setCell(table,6,3,c.yiqibh) #setCell(table,6,3).Value =c.yiqibh
    setCell(table,3,3,c.yonghu) #setCell(table,3,3).Value =c.yonghu
    d=datetime.datetime.now()
    setCell(table,13,4,str(d.year)+"年") #setCell(table,13,4).FormulaR1C1 =str(d.year)+"年"
    setCell(table,13,5,str(d.month)+"月")#setCell(table,13,5).Value =str(d.month)+"月"
    setCell(table,13,6,str(d.day)+"日")#setCell(table,13,6).Value =str(d.day)+"日"    
    setCell(table,14,3,c.baoxiang)#setCell(table,13,6).Value =str(d.day)+"日"    
    return save_virtual_workbook(xlBook)
def genJiaozhunONH(c,fn):
    xlBook = load_workbook(filename = fn)
    table=xlBook.worksheets[0] 
    setCell(table,10,8,c.yiqixinghao)# setCell(table,11,8).Value =c.yiqixinghao
    setCell(table,16,8,c.yiqibh)# setCell(table,16,8).Value =c.yiqibh
    setCell(table,18,8,c.yonghu)# setCell(table,18,8).Value =c.yonghu
    setCell(table,20,8,c.addr)# setCell(table,18,8).Value =c.yonghu
    d=c.yujifahuo_date
    d1=d+datetime.timedelta(-1)
    setCell(table,32,8,str(d1.year))# setCell(table,32,8).FormulaR1C1 =d1.year#年
    setCell(table,32,12,str(d1.month))# setCell(table,32,12).Value =d1.month#月
    setCell(table,32,16,str(d1.day))# setCell(table,32,16).Value =d1.day#日
    d2=d+datetime.timedelta(364)
    setCell(table,35,8,str(d2.year))# setCell(table,35,8).Value =d2.year#年
    setCell(table,35,12,str(d2.month))# setCell(table,35,12).Value =d2.month#月
    setCell(table,35,16,str(d2.day))# setCell(table,35,16).Value =d2.day#日
    # #page 2
    table=xlBook.worksheets[1] 
    dd="  地点（LOCATION）： "+c.yonghu
    setCell(table,23,3,"地点"+"("+"LOCATION"+")"+":"+"    "+c.yonghu)# setCell(table,22,3).Value=ddsetCell(table,22,3,dd)# setCell(table,26,3,dd
    # #page 3
    table=xlBook.worksheets[2] 
    eles=[]
    for i in [6,9,12,14,17,20]:
        eles.append(getCell(table,13,i))
    stds=[]
    for i in [6,9,12,14,17,20]:
        stds.append(getCell(table,14,i))
    # print(eles,stds)
    (tests,errs)=genTestR(eles,stds)
    tmp=0
    for i in [6,9,12,14,17,20]:
        setCell(table,15,i,tests[tmp])
        tmp +=1
    tmp=0
    for i in [6,9,12,14,17,20]:
        setCell(table,16,i,errs[tmp])
        tmp +=1
    #jmd
    cave=0.0112
    crsd=0.8/100
    (rs,ave_str,rsd_str)=genjmd(cave,crsd)
    cjmd_str=",".join(rs)
    setCell(table,20,3,"   测量值(O/%):" +cjmd_str)
    setCell(table,21,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    sjmd_str=" 0.0721,0.0725,0.0723,0.0727,0.0730,0.0725,0.0725"
    save=0.0084
    srsd=1.3/100
    (rs,ave_str,rsd_str)=genjmd(save,srsd)
    sjmd_str=",".join(rs)
    setCell(table,22,3,"   测量值(N/%):" + sjmd_str)
    setCell(table,23,3,"   平均值:%s,   相对标准偏差:%s" % (ave_str+"%",rsd_str+"%"))
    return save_virtual_workbook(xlBook)

def getJiaoZhunFile(c):
    if "-" in c.yiqixinghao:
        (lx,tmp)=c.yiqixinghao.split("-")
        print(lx)
    else:
        lx=c.yiqixinghao
    data=None
    if lx==u"O":
        tname="O模板"
        fullfilepath = os.path.join(MEDIA_ROOT,"t_"+tname+".xlsx")
        logging.info(fullfilepath)
        data=genJiaozhunO(c,fullfilepath)
    elif lx==u"N":
        tname="N模板"
        fullfilepath = os.path.join(MEDIA_ROOT,"t_"+tname+".xlsx")
        logging.info(fullfilepath)
        data=genJiaozhunN(c,fullfilepath)
    elif lx==u"ON":
        tname="ON模板"
        fullfilepath = os.path.join(MEDIA_ROOT,"t_"+tname+".xlsx")
        logging.info(fullfilepath)
        data=genJiaozhunON(c,fullfilepath)
    elif lx==u"OH":
        tname="OH模板"
        fullfilepath = os.path.join(MEDIA_ROOT,"t_"+tname+".xlsx")
        logging.info(fullfilepath)
        data=genJiaozhunOH(c,fullfilepath)
    elif lx==u"ONH":
        tname="ONH模板"
        fullfilepath = os.path.join(MEDIA_ROOT,"t_"+tname+".xlsx")
        logging.info(fullfilepath)
        data=genJiaozhunONH(c,fullfilepath)
    else:
        tname="CS模板"#"aveSingle2"
        fullfilepath = os.path.join(MEDIA_ROOT,"t_"+tname+".xlsx")
        logging.info(fullfilepath)
        data=genJiaozhunCS(c,fullfilepath)
    return data
if __name__=="__main__":
    print(genPack("4111533499"))